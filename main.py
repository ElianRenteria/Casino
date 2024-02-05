import pygame
import random, math
from slot import *
from button import Button
from openpyxl import Workbook, load_workbook

"""workbook = Workbook()
spreadsheet = workbook.active

spreadsheet["users"] = "Hello"
workbook.save(filename="HelloWorld.xlsx")
workbook = load_workbook(filename="HelloWorld.xlsx")
print(workbook.sheetnames)

spreadsheet = workbook.active
print(spreadsheet)"""


pygame.init()

spinning_sound = pygame.mixer.Sound("sounds/spinning.mp3")
spinning_sound.fadeout(25)
spinning_sound.set_volume(.5)
smallWin_sound = pygame.mixer.Sound("sounds/smallWin2.mp3")
mediumWin_sound = pygame.mixer.Sound("sounds/mediumWin.wav")
bigWin_sound = pygame.mixer.Sound("sounds/bigWin.wav")
bg_sound = pygame.mixer.Sound("sounds/backgroundMusic.mp3")
noWin_sound = pygame.mixer.Sound("sounds/noWin.mp3")
click_sound = pygame.mixer.Sound("sounds/button.mp3")


font = pygame.font.Font('freesansbold.ttf', 48)
font2 = pygame.font.Font('freesansbold.ttf', 24)

pygame.display.set_caption('Fortune Time')
window = pygame.display.set_mode((1000, 1000))
board = [Slot(0, 0), Slot(200, 0), Slot(400, 0), Slot(600, 0), Slot(800, 0), Slot(0, 800), Slot(200, 800), Slot(400, 800), Slot(600, 800), Slot(800, 800), Slot(0, 200), Slot(0, 400), Slot(0, 600), Slot(800, 200), Slot(800, 400), Slot(800, 600)]
board[random.randint(0,2)].reload(2)
board[random.randint(3,5)].reload(4)
board[random.randint(6,9)].reload(1)
board[random.randint(10,12)].reload(3)
board[random.randint(13,15)].reload(0)
money = 100
buttons = []
colors = [(220, 20, 60), (255, 215, 0), (51, 204, 51), (255, 127, 127), (135, 31, 120)]
types = ["pineapple", "watermelon", "apple", "grape"]
start_button = Button((0, 100, 204), window, 0)
start_button.set_x(575)
start_button.set_y(475)
start_button.set_rect(200,100)
start_button.bet = -3
for i in range(0, 5):
    buttons.append(Button(colors[i], window, i))
rates = {"lemon":0,}
spin_amount = -1
flashing = False
def set_rates():
    global types, board
    for type in types:
        count = 0
        for slot in board:
            if slot.type == type:
                count += 1
        if count > 0:
            rates[type] = (math.ceil(1/(count/16)))
        else:
            rates[type] = 0
set_rates()
count = 1
buttons[0].multiplier = 0
for type in types:
    buttons[count].multiplier = rates[type]
    count += 1
def payout(winner_type):
    global money, buttons
    won = False
    if winner_type == "pineapple":
        if buttons[1].bet > 0:
            money += rates[winner_type] * buttons[1].bet
            won = True
    elif winner_type == "watermelon":
        if buttons[2].bet > 0:
            money += rates[winner_type] * buttons[2].bet
            won = True
    elif winner_type == "apple":
        if buttons[3].bet > 0:
            money += rates[winner_type] * buttons[3].bet
            won = True
    elif winner_type == "grape":
        if buttons[4].bet > 0:
            money += rates[winner_type] * buttons[4].bet
            won = True
    if rates[winner_type] and won:
        if rates[winner_type] >= 10:
            bigWin_sound.play(0)
        elif rates[winner_type] >= 5:
            mediumWin_sound.play(0)
        else:
            smallWin_sound.play(0)
    else:
        noWin_sound.play(0)
    for i in range(0, 5):
        buttons[i].bet = 0
    if money <= 0:
        print("no more money")


def flash(x, y):
    global flashing
    for i in range(0,20):
        if flashing:
            window.fill((255, 255, 255))
            window.blit(pygame.transform.scale(pygame.image.load("assets/bg.jpg"),(600,600)),(200, 200))
            for slot in board:
                slot.draw(window)
            window.blit(pygame.transform.scale(pygame.image.load("assets/logo2.png"), (350, 200)), (325, 225))
            for button in buttons:
                button.draw()
            start_button.draw()
            transparent_surface = pygame.Surface((200, 200), pygame.SRCALPHA)
            transparent_surface.fill((0, 0, 255, 150))  # Set alpha value to 128 for semi-transparency
            window.blit(transparent_surface, (x, y))
            balance_text = font2.render("Balance:", True, (40, 40, 40))
            money_text = font.render("$" + str(money), True, (40, 40, 40))
            window.blit(balance_text, (250, 450))
            window.blit(money_text, (250, 500))
            pygame.display.update()
        else:
            window.fill((255, 255, 255))
            #pygame.draw.rect(window, (192, 192, 192), pygame.Rect(200, 200, 600, 600))
            window.blit(pygame.transform.scale(pygame.image.load("assets/bg.jpg"),(600,600)),(200, 200))
            for slot in board:
                slot.draw(window)
            window.blit(pygame.transform.scale(pygame.image.load("assets/logo2.png"), (350, 200)), (325, 225))
            for button in buttons:
                button.draw()
            start_button.draw()
            balance_text = font2.render("Balance:", True, (40, 40, 40))
            money_text = font.render("$" + str(money), True, (40, 40, 40))
            window.blit(balance_text, (250, 450))
            window.blit(money_text, (250, 500))
            pygame.display.update()
        flashing = not flashing
        pygame.time.wait(100)
def spin():
    global spin_amount
    chosen_spot = board[random.randint(0,15)]
    transparent_surface = pygame.Surface((200, 200), pygame.SRCALPHA)
    transparent_surface.fill((0, 0, 255, 150))  # Set alpha value to 128 for semi-transparency
    window.blit(transparent_surface, (chosen_spot.x, chosen_spot.y))
    balance_text = font2.render("Balance:", True, (40, 40, 40))
    money_text = font.render("$" + str(money), True, (40, 40, 40))
    window.blit(balance_text, (250, 450))
    window.blit(money_text, (250, 500))
    pygame.display.update()
    spinning_sound.play(0,300)
    pygame.time.wait(150)
    spin_amount -= 1
    if spin_amount <= 0:
        pygame.time.wait(100)
        payout(chosen_spot.type)
        flash(chosen_spot.x, chosen_spot.y)
        start_button.active = True
        for button in buttons:
            button.active = True

def draw():
    global spin_amount, money
    window.fill((255, 255, 255))
    window.blit(pygame.transform.scale(pygame.image.load("assets/bg.jpg"),(600,600)),(200, 200))
    for slot in board:
        slot.draw(window)
    window.blit(pygame.transform.scale(pygame.image.load("assets/logo2.png"),(350, 200)), (325, 225))
    for button in buttons:
        button.draw()
        if button.pressed:
            if money > 0:
                click_sound.play(0,0,20)
                button.bet += 1
                money -= 1
                pygame.time.wait(150)
            button.pressed = False
    start_button.draw()
    if start_button.pressed:
        start_button.pressed = False
        start_button.active = False
        for button in buttons:
            button.active = False
        spin_amount = random.randint(10, 15)
    if spin_amount > 0:
        spin()
    balance_text = font2.render("Balance:", True, (40, 40, 40))
    money_text = font.render("$" + str(money), True, (40, 40, 40))
    window.blit(balance_text, (250, 450))
    window.blit(money_text, (250,500))

run = True
clock = pygame.time.Clock()
pygame.mixer.music.load("sounds/backgroundMusic.mp3")
pygame.mixer.music.play(-1)
pygame.mixer.music.set_volume(.8)
while run:
    clock.tick(60)
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False
    draw()
    pygame.display.update()
pygame.quit()
quit()








